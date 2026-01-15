from typing import List
import io
import logging
import uuid
from sqlalchemy.orm import Session

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from src.models import JiraTicket, Invoice, InvoiceLine, Currency
from src.database import InvoiceRepository, ClauseRepository
from src.database.models import CurrencyEnum, InvoiceStatusEnum
from .mapping_engine import MappingEngine

# Configure logging
logger = logging.getLogger(__name__)


class InvoiceGenerator:
    """Service for generating and exporting invoices"""
    
    @staticmethod
    def generate_from_tickets(
        project_name: str,
        billing_period: str,
        tickets: List[JiraTicket],
        created_by: str,
        db: Session
    ) -> Invoice:
        """
        Generate an invoice from a list of enriched tickets
        Tickets must already be enriched with billing information
        """
        logger.info(f"=== InvoiceGenerator.generate_from_tickets ===")
        logger.info(f"Input: {len(tickets)} tickets")
        
        # Process tickets to validate and aggregate
        mapping_results = MappingEngine.process_tickets_batch(tickets, db)
        logger.info(f"Mapping results: {len(mapping_results['valid'])} valid, {len(mapping_results['invalid'])} invalid")
        
        if not mapping_results["valid"]:
            logger.error("No valid billable tickets after processing!")
            for invalid in mapping_results["invalid"]:
                logger.error(f" Invalid: {invalid['ticket'].ticket_id} - {invalid['error']}")
            raise ValueError("No valid billable tickets found")
        
        # Create invoice lines from valid tickets
        lines = []
        for item in mapping_results["valid"]:
            ticket = item["ticket"]
            clause = item["clause"]
            
            logger.info(f"Creating line for {ticket.ticket_id}: {ticket.hours_worked}h × €{clause.unit_price} = €{item['cost']}")
            
            line_data = {
                "jira_ticket_id": ticket.ticket_id,
                "clause_id": ticket.clause_id,
                "hours_worked": float(ticket.hours_worked),
                "unit_price": float(clause.unit_price),
                "line_total": float(item['cost'])
            }
            lines.append(line_data)
        
        # Generate unique invoice ID
        invoice_id = str(uuid.uuid4())

        # Create invoice
        invoice_data = {
            "invoice_id": invoice_id,
            "project_name": project_name,
            "billing_period": billing_period,
            "total_amount": float(mapping_results["total_cost"]),
            "currency": CurrencyEnum.EUR,
            "status": InvoiceStatusEnum.DRAFT,
            "created_by": created_by
        }
        
        logger.info(f"Creating invoice with total: €{mapping_results['total_cost']}")
        
        # Use repository to create invoice
        invoice_repo = InvoiceRepository(db)
        invoice_model = invoice_repo.create_with_lines(invoice_data, lines)
        
        logger.info(f"Invoice created: {invoice_model.invoice_id}")
        
        # Convert to Pydantic model
        invoice = InvoiceGenerator._convert_invoice_model(invoice_model, db)
        return invoice
    
    @staticmethod
    def _convert_invoice_model(invoice_model, db: Session) -> Invoice:
        """Convert SQLAlchemy InvoiceModel to Pydantic Invoice"""
        clause_repo = ClauseRepository(db)
        
        lines = []
        for line_model in invoice_model.lines:
            line = InvoiceLine(
                line_id=line_model.line_id,
                invoice_id=line_model.invoice_id,
                jira_ticket_id=line_model.jira_ticket_id,
                clause_id=line_model.clause_id,
                hours_worked=line_model.hours_worked,
                unit_price=line_model.unit_price,
                line_total=line_model.line_total
            )
            lines.append(line)
        
        return Invoice(
            invoice_id=invoice_model.invoice_id,
            project_name=invoice_model.project_name,
            billing_period=invoice_model.billing_period,
            total_amount=invoice_model.total_amount,
            currency=Currency(invoice_model.currency.value),
            status=invoice_model.status.value,
            created_by=invoice_model.created_by,
            created_at=invoice_model.created_at,
            lines=lines
        )
    
    @staticmethod
    def export_to_pdf(invoice: Invoice, db: Session) -> bytes:
        """Export invoice to PDF format"""
        buffer = io.BytesIO()
        clause_repo = ClauseRepository(db)
        
        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph(
            f"INVOICE {invoice.invoice_id}",
            styles['Title']
        )
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Invoice details
        details_text = f"""
        Project: {invoice.project_name}
        Billing Period: {invoice.billing_period}
        Date: {invoice.created_at.strftime('%Y-%m-%d')}
        Status: {invoice.status.value}
        """
        details = Paragraph(details_text, styles['Normal'])
        elements.append(details)
        elements.append(Spacer(1, 20))
        
        # Line items table
        table_data = [
            ['Ticket ID', 'Clause', 'Hours', 'Rate', 'Total']
        ]
        
        for line in invoice.lines:
            clause = clause_repo.get(line.clause_id)
            table_data.append([
                line.jira_ticket_id,
                f"{clause.clause_name}" if clause else line.clause_id,
                f"{line.hours_worked}",
                f"{line.unit_price} {invoice.currency.value}",
                f"{line.line_total} {invoice.currency.value}"
            ])
        
        # Add total row
        table_data.append([
            '', '', '', 'TOTAL:',
            f"{invoice.total_amount} {invoice.currency.value}"
        ])
        
        # Create table
        table = Table(table_data, colWidths=[80, 150, 60, 80, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        return buffer.getvalue()
    
    @staticmethod
    def export_to_excel(invoice: Invoice, db: Session) -> bytes:
        """Export invoice to Excel format (BMW SAP compatible)"""
        clause_repo = ClauseRepository(db)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Invoice header
        ws['A1'] = 'Invoice ID'
        ws['B1'] = invoice.invoice_id
        ws['A2'] = 'Project'
        ws['B2'] = invoice.project_name
        ws['A3'] = 'Billing Period'
        ws['B3'] = invoice.billing_period
        ws['A4'] = 'Date'
        ws['B4'] = invoice.created_at.strftime('%Y-%m-%d')
        
        # Line items header
        headers = ['Ticket ID', 'Clause ID', 'Clause Name', 'Hours', 'Unit Price', 'Currency', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Line items
        row = 7
        for line in invoice.lines:
            clause = clause_repo.get(line.clause_id)
            ws.cell(row=row, column=1).value = line.jira_ticket_id
            ws.cell(row=row, column=2).value = line.clause_id
            ws.cell(row=row, column=3).value = clause.clause_name if clause else ''
            ws.cell(row=row, column=4).value = float(line.hours_worked)
            ws.cell(row=row, column=5).value = float(line.unit_price)
            ws.cell(row=row, column=6).value = invoice.currency.value
            ws.cell(row=row, column=7).value = float(line.line_total)
            row += 1
        
        # Total row
        total_row = row
        ws.cell(row=total_row, column=6).value = 'TOTAL:'
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=7).value = float(invoice.total_amount)
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        
        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    @staticmethod
    def export_to_sap_xml(invoice: Invoice, db: Session) -> str:
        """Export invoice to SAP XML format"""
        clause_repo = ClauseRepository(db)
        
        xml_lines = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            '<Invoice>',
            f'  <InvoiceID>{invoice.invoice_id}</InvoiceID>',
            f'  <ProjectName>{invoice.project_name}</ProjectName>',
            f'  <BillingPeriod>{invoice.billing_period}</BillingPeriod>',
            f'  <TotalAmount>{invoice.total_amount}</TotalAmount>',
            f'  <Status>{invoice.status.value}</Status>',
            f'  <CreatedAt>{invoice.created_at.isoformat()}</CreatedAt>',
            '  <Lines>'
        ]
        
        for line in invoice.lines:
            clause = clause_repo.get(line.clause_id)
            xml_lines.extend([
                '    <Line>',
                f'      <TicketID>{line.jira_ticket_id}</TicketID>',
                f'      <ClauseID>{line.clause_id}</ClauseID>',
                f'      <ClauseName>{clause.clause_name if clause else ""}</ClauseName>',
                f'      <Hours>{line.hours_worked}</Hours>',
                f'      <UnitPrice>{line.unit_price}</UnitPrice>',
                f'      <Total>{line.line_total}</Total>',
                '    </Line>'
            ])
        
        xml_lines.extend([
            '  </Lines>',
            '</Invoice>'
        ])
        
        return '\n'.join(xml_lines)