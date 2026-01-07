from typing import List, Dict, Optional, Tuple
from decimal import Decimal
from datetime import datetime, timezone
import io
import logging
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from models import (
    JiraTicket, LegalClause, Invoice, InvoiceLine,
    TicketStatus, Currency
)
from database import db

# Configure logging
logger = logging.getLogger(__name__)


# ============================================================================
# JIRA INTEGRATION SERVICE
# ============================================================================

class JiraIntegrationService:
    """Service for interacting with Jira API and enriching tickets"""
    
    def __init__(self, api_endpoint: str, user_token: str):
        self.api_endpoint = api_endpoint
        self.user_token = user_token
    
    async def fetch_and_process_tickets(
        self,
        project_key: str,
        status_filter: Optional[TicketStatus] = None,
        label_filter: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> List[JiraTicket]:
        """
        Fetch tickets from Jira API and enrich them with billing information
        This is the SINGLE source of truth for fetching tickets
        """
        # Step 1: Fetch raw tickets from database (in production, from Jira API)
        tickets = db.get_jira_tickets(status=status_filter, label=label_filter)
        
        # Step 2: Filter by date range if provided
        if start_date or end_date:
            tickets = self._filter_by_date_range(tickets, start_date, end_date)
        
        # Step 3: Enrich each ticket with billing information
        enriched_tickets = []
        for ticket in tickets:
            enriched_ticket = self._enrich_ticket_with_billing(ticket)
            enriched_tickets.append(enriched_ticket)
        
        return enriched_tickets
    
    def _normalize_datetime(self, dt: Optional[datetime]) -> Optional[datetime]:
        """
        Normalize datetime to be timezone-aware (UTC)
        If datetime is naive, assume it's UTC
        """
        if dt is None:
            return None
        
        if dt.tzinfo is None:
            # Naive datetime - assume UTC
            return dt.replace(tzinfo=timezone.utc)
        else:
            # Already timezone-aware - convert to UTC
            return dt.astimezone(timezone.utc)
    
    def _filter_by_date_range(
        self,
        tickets: List[JiraTicket],
        start_date: Optional[datetime],
        end_date: Optional[datetime]
    ) -> List[JiraTicket]:
        """Filter tickets by date range"""
        filtered = tickets
        
        # Normalize input dates to UTC
        start_date_utc = self._normalize_datetime(start_date)
        end_date_utc = self._normalize_datetime(end_date)
        
        if start_date_utc:
            filtered = [
                t for t in filtered 
                if t.resolved_at and self._normalize_datetime(t.resolved_at) >= start_date_utc
            ]
        
        if end_date_utc:
            filtered = [
                t for t in filtered 
                if t.resolved_at and self._normalize_datetime(t.resolved_at) <= end_date_utc
            ]
        
        return filtered
    
    def _enrich_ticket_with_billing(self, ticket: JiraTicket) -> JiraTicket:
        """
        Enrich a single ticket with billing information
        This modifies the ticket object in place
        """
        # Match ticket to clause
        clause = MappingEngine.match_ticket_to_clause(ticket)
        
        if clause:
            # Calculate billable amount
            billable_amount = MappingEngine.calculate_line_cost(ticket, clause)
            
            # Update ticket with billing info
            ticket.clause_id = clause.clause_id
            ticket.billable_amount = billable_amount
            ticket.is_billable = True
            
            print(f"Enriched Ticket {ticket.ticket_id}: Clause={clause.clause_id}, Amount={billable_amount}")
        else:
            # No matching clause found
            ticket.clause_id = None
            ticket.billable_amount = Decimal("0.00")
            ticket.is_billable = False
            
            print(f"Ticket {ticket.ticket_id}: No matching clause found")
        
        return ticket
    
    async def authenticate(self) -> bool:
        """Validate Jira token"""
        # In production, make a test API call to validate token
        return bool(self.user_token)


# ============================================================================
# MAPPING ENGINE SERVICE
# ============================================================================

class MappingEngine:
    """Service for mapping Jira tickets to legal clauses"""
    
    @staticmethod
    def match_ticket_to_clause(ticket: JiraTicket) -> Optional[LegalClause]:
        """
        Match a Jira ticket to a legal clause based on labels
        Returns the first matching active clause or None
        """
        if not ticket.labels:
            return None
        
        # Try to find a matching clause for any label
        for label in ticket.labels:
            clause = db.get_clause(label)
            if clause and clause.is_active:
                return clause
        
        return None
    
    @staticmethod
    def calculate_line_cost(ticket: JiraTicket, clause: LegalClause) -> Decimal:
        """Calculate the cost for a ticket based on hours and clause price"""
        if not ticket.hours_worked or not clause.unit_price:
            return Decimal("0.00")
        
        cost = ticket.hours_worked * clause.unit_price
        return round(cost, 2)
    
    @staticmethod
    def validate_mapping(ticket: JiraTicket) -> Tuple[bool, str]:
        """
        Validate if a ticket can be mapped and billed
        Returns (is_valid, error_message)
        """
        if ticket.status != TicketStatus.CLOSED:
            return False, f"Ticket {ticket.ticket_id} is not closed"
        
        if not ticket.labels:
            return False, f"Ticket {ticket.ticket_id} has no labels"
        
        if not ticket.hours_worked or ticket.hours_worked <= 0:
            return False, f"Ticket {ticket.ticket_id} has no hours logged"
        
        # Check if any label matches a clause
        clause = MappingEngine.match_ticket_to_clause(ticket)
        if not clause:
            return False, f"Ticket {ticket.ticket_id} has no matching legal clause"
        
        return True, ""
    
    @staticmethod
    def process_tickets_batch(tickets: List[JiraTicket]) -> Dict:
        """
        Process multiple tickets and return mapping results
        Assumes tickets are already enriched with billing info
        """
        results = {
            "valid": [],
            "invalid": [],
            "total_cost": Decimal("0.00"),
            "total_hours": Decimal("0.00")
        }
        
        for ticket in tickets:
            is_valid, error = MappingEngine.validate_mapping(ticket)
            
            if is_valid and ticket.is_billable:
                clause = db.get_clause(ticket.clause_id)
                
                results["valid"].append({
                    "ticket": ticket,
                    "clause": clause,
                    "cost": ticket.billable_amount
                })
                results["total_cost"] += ticket.billable_amount
                results["total_hours"] += ticket.hours_worked
            else:
                results["invalid"].append({
                    "ticket": ticket,
                    "error": error if error else "Not billable"
                })
        
        return results


# ============================================================================
# INVOICE GENERATOR SERVICE
# ============================================================================

class InvoiceGenerator:
    """Service for generating and exporting invoices"""
    
    @staticmethod
    def generate_from_tickets(
        project_name: str,
        billing_period: str,
        tickets: List[JiraTicket],
        created_by: str
    ) -> Invoice:
        """
        Generate an invoice from a list of enriched tickets
        Tickets must already be enriched with billing information
        """
        logger.info(f"=== InvoiceGenerator.generate_from_tickets ===")
        logger.info(f"Input: {len(tickets)} tickets")
        
        # Process tickets to validate and aggregate
        mapping_results = MappingEngine.process_tickets_batch(tickets)
        
        logger.info(f"Mapping results: {len(mapping_results['valid'])} valid, {len(mapping_results['invalid'])} invalid")
        
        if not mapping_results["valid"]:
            logger.error("No valid billable tickets after processing!")
            for invalid in mapping_results["invalid"]:
                logger.error(f"  Invalid: {invalid['ticket'].ticket_id} - {invalid['error']}")
            raise ValueError("No valid billable tickets found")
        
        # Create invoice lines from valid tickets
        lines = []
        for item in mapping_results["valid"]:
            ticket = item["ticket"]
            clause = item["clause"]
            
            logger.info(f"Creating line for {ticket.ticket_id}: {ticket.hours_worked}h × €{clause.unit_price} = €{item['cost']}")
            
            line_data = {
                "jira_ticket_id": ticket.ticket_id,
                "clause_id": ticket.clause_id,
                "hours_worked": ticket.hours_worked,
                "unit_price": clause.unit_price
            }
            lines.append(line_data)
        
        # Create invoice
        invoice_data = {
            "project_name": project_name,
            "billing_period": billing_period,
            "total_amount": mapping_results["total_cost"],
            "currency": Currency.EUR
        }
        
        logger.info(f"Creating invoice with total: €{mapping_results['total_cost']}")
        
        invoice = db.create_invoice(
            invoice_data=invoice_data,
            lines=lines,
            created_by=created_by
        )
        
        logger.info(f"Invoice created: {invoice.invoice_id}")
        
        return invoice
    
    @staticmethod
    def export_to_pdf(invoice: Invoice) -> bytes:
        """Export invoice to PDF format"""
        buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph(
            f"<b>INVOICE {invoice.invoice_id}</b>",
            styles['Title']
        )
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Invoice details
        details_text = f"""
        <b>Project:</b> {invoice.project_name}<br/>
        <b>Billing Period:</b> {invoice.billing_period}<br/>
        <b>Date:</b> {invoice.created_at.strftime('%Y-%m-%d')}<br/>
        <b>Status:</b> {invoice.status.value}<br/>
        """
        details = Paragraph(details_text, styles['Normal'])
        elements.append(details)
        elements.append(Spacer(1, 20))
        
        # Line items table
        table_data = [
            ['Ticket ID', 'Clause', 'Hours', 'Rate', 'Total']
        ]
        
        for line in invoice.lines:
            clause = db.get_clause(line.clause_id)
            table_data.append([
                line.jira_ticket_id,
                f"{clause.clause_name}" if clause else line.clause_id,
                f"{line.hours_worked}",
                f"{line.unit_price} {invoice.currency.value}",
                f"{line.line_total} {invoice.currency.value}"
            ])
        
        # Add total row
        table_data.append([
            '', '', '', '<b>TOTAL:</b>',
            f"<b>{invoice.total_amount} {invoice.currency.value}</b>"
        ])
        
        # Create table
        table = Table(table_data, colWidths=[80, 150, 60, 80, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        return buffer.getvalue()
    
    @staticmethod
    def export_to_excel(invoice: Invoice) -> bytes:
        """Export invoice to Excel format (BMW SAP compatible)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Invoice header
        ws['A1'] = 'Invoice ID'
        ws['B1'] = invoice.invoice_id
        ws['A2'] = 'Project'
        ws['B2'] = invoice.project_name
        ws['A3'] = 'Billing Period'
        ws['B3'] = invoice.billing_period
        ws['A4'] = 'Date'
        ws['B4'] = invoice.created_at.strftime('%Y-%m-%d')
        
        # Line items header
        headers = ['Ticket ID', 'Clause ID', 'Clause Name', 'Hours', 'Unit Price', 'Currency', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Line items
        row = 7
        for line in invoice.lines:
            clause = db.get_clause(line.clause_id)
            ws.cell(row=row, column=1).value = line.jira_ticket_id
            ws.cell(row=row, column=2).value = line.clause_id
            ws.cell(row=row, column=3).value = clause.clause_name if clause else ''
            ws.cell(row=row, column=4).value = float(line.hours_worked)
            ws.cell(row=row, column=5).value = float(line.unit_price)
            ws.cell(row=row, column=6).value = invoice.currency.value
            ws.cell(row=row, column=7).value = float(line.line_total)
            row += 1
        
        # Total row
        total_row = row
        ws.cell(row=total_row, column=6).value = 'TOTAL:'
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=7).value = float(invoice.total_amount)
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        
        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    @staticmethod
    def export_to_sap_xml(invoice: Invoice) -> str:
        """Export invoice to SAP XML format"""
        xml_lines = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            '<Invoice>',
            f'  <InvoiceID>{invoice.invoice_id}</InvoiceID>',
            f'  <ProjectName>{invoice.project_name}</ProjectName>',
            f'  <BillingPeriod>{invoice.billing_period}</BillingPeriod>',
            f'  <TotalAmount currency="{invoice.currency.value}">{invoice.total_amount}</TotalAmount>',
            f'  <Status>{invoice.status.value}</Status>',
            f'  <CreatedDate>{invoice.created_at.isoformat()}</CreatedDate>',
            '  <LineItems>'
        ]
        
        for line in invoice.lines:
            clause = db.get_clause(line.clause_id)
            xml_lines.extend([
                '    <LineItem>',
                f'      <TicketID>{line.jira_ticket_id}</TicketID>',
                f'      <ClauseID>{line.clause_id}</ClauseID>',
                f'      <ClauseName>{clause.clause_name if clause else ""}</ClauseName>',
                f'      <Hours>{line.hours_worked}</Hours>',
                f'      <UnitPrice>{line.unit_price}</UnitPrice>',
                f'      <LineTotal>{line.line_total}</LineTotal>',
                '    </LineItem>'
            ])
        
        xml_lines.extend([
            '  </LineItems>',
            '</Invoice>'
        ])
        
        return '\n'.join(xml_lines)


# ============================================================================
# ANALYTICS SERVICE
# ============================================================================

class AnalyticsService:
    """Service for generating reports and analytics"""
    
    @staticmethod
    def get_monthly_summary(billing_period: str, user_id: Optional[str] = None) -> Dict:
        """Generate monthly billing summary"""
        invoices = db.list_invoices(created_by=user_id)
        
        # Filter by billing period
        period_invoices = [
            inv for inv in invoices
            if inv.billing_period == billing_period
        ]
        
        total_amount = sum(inv.total_amount for inv in period_invoices)
        total_hours = sum(
            sum(line.hours_worked for line in inv.lines)
            for inv in period_invoices
        )
        total_tickets = sum(len(inv.lines) for inv in period_invoices)
        
        # Breakdown by clause
        clause_breakdown = {}
        for invoice in period_invoices:
            for line in invoice.lines:
                clause_id = line.clause_id
                if clause_id not in clause_breakdown:
                    clause = db.get_clause(clause_id)
                    clause_breakdown[clause_id] = {
                        "clause_name": clause.clause_name if clause else clause_id,
                        "hours": Decimal("0.00"),
                        "amount": Decimal("0.00"),
                        "tickets": 0
                    }
                
                clause_breakdown[clause_id]["hours"] += line.hours_worked
                clause_breakdown[clause_id]["amount"] += line.line_total
                clause_breakdown[clause_id]["tickets"] += 1
        
        return {
            "billing_period": billing_period,
            "total_hours": float(total_hours),
            "total_amount": float(total_amount),
            "tickets_billed": total_tickets,
            "invoices_count": len(period_invoices),
            "breakdown_by_clause": clause_breakdown
        }
    
    @staticmethod
    def get_user_performance(user_id: str, months: int = 3) -> Dict:
        """Get performance metrics for a user"""
        invoices = db.list_invoices(created_by=user_id)
        
        # Calculate metrics
        total_invoices = len(invoices)
        total_revenue = sum(inv.total_amount for inv in invoices)
        
        # Status breakdown
        status_count = {}
        for invoice in invoices:
            status = invoice.status.value
            status_count[status] = status_count.get(status, 0) + 1
        
        return {
            "user_id": user_id,
            "total_invoices": total_invoices,
            "total_revenue": float(total_revenue),
            "status_breakdown": status_count,
            "avg_invoice_amount": float(total_revenue / total_invoices) if total_invoices > 0 else 0
        }